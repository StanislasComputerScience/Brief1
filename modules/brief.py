import json
import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from typing import Any, Optional, Tuple, Dict, List
import os

def importation_des_donnees(fichier: str) -> dict:
    """Charge un fichier JSON et retourne son contenu sous forme de dictionnaire."""
    chemin = os.path.join(os.path.dirname(__file__), '..', fichier)
    chemin = os.path.abspath(chemin)  # chemin absolu
    with open(chemin, encoding='utf-8') as file:
        return json.load(file)

def calcul_salaire(**employe) -> Optional[Tuple[float, int]]:
    """Calcule le salaire mensuel d'un employé et retourne un tuple (salaire, option)
    Où option est égale à : heures_supplementaires = heures_travaillees - heures_contractuelles.

    NB: On a fait le choix de ne payer que les heures travaillées et non les heures prévues.

    On a considéré que le salaire mensuel était le quadruple du salaire hebdomadaire.
    """
    if not employe:
        return None

    taux_horaire = employe['hourly_rate']
    heures_contractuelles = employe['contract_hours']
    heures_travaillees = employe['weekly_hours_worked']
    heures_supplementaires = heures_travaillees - heures_contractuelles

    if heures_supplementaires == 0:
        salaire_hebdo = heures_travaillees * taux_horaire
        return salaire_hebdo * 4, 0
    elif heures_supplementaires > 0:
        salaire_hebdo = (heures_contractuelles + heures_supplementaires * 1.5) * taux_horaire
        return salaire_hebdo * 4, heures_supplementaires
    else:
        salaire_hebdo = heures_travaillees * taux_horaire
        return salaire_hebdo * 4, heures_supplementaires

def statistiques_filiales(donnees: dict) -> List[Dict[str, Any]]:
    """ Génère les statistiques pour chaque filiale à partir d'un dictionnaire
        représantant une filiale.
        Pour la filiale entrée dans donnée, on collecte le nom de la filiale puis,
        en parcourant tous ses employes (i.e. employes) dont on calcul le salaire,
        on calcul le plus grand salaire (=salaireMax)
        on calcul le plus petit salaire (=salaireMin)

        on calcul le nombre d'employes  (=effectif)
        on calcul la somme de tous les salaires (=somme_salaire)
        ces deux derniers permettent de calculer le salaire moyen (=salaireMoyen)
        et seront utiles tout à l'heure pour calculer le salaire moyen de toute
        l'entreprise.
        Ces informations sont collectées dans le dictionnaire info_filiale.
        Enfin on concatène tous ces dictionnaires dans la liste stats,
        que l'on retourne à la fin.
    """
    stats = []
    for filiale, employes in donnees.items():
        info_filiale = {
            "nom": filiale,
            "salaireMax": float('-inf'),
            "salaireMin": float('inf'),
            "effectif": 0,
            "somme_salaire": 0,
            "listeEmployes": []
        }

        for employe in employes:
            salaire, option = calcul_salaire(**employe)
            employe.update({
                "salary": salaire,
                "info": option
            })
            info_filiale["salaireMax"] = max(info_filiale["salaireMax"], salaire)
            info_filiale["salaireMin"] = min(info_filiale["salaireMin"], salaire)
            info_filiale["effectif"] += 1
            info_filiale["somme_salaire"] += salaire
            info_filiale["listeEmployes"].append(employe)

        if info_filiale["effectif"] > 0:
            info_filiale["salaireMoyen"] = info_filiale["somme_salaire"] / info_filiale["effectif"]
        else:
            info_filiale["salaireMoyen"] = 0

        stats.append(info_filiale)

    return stats

def affichage(stats: List[Dict[str, Any]]) -> str:
    """
    A partir de la liste retournée par la fonction statistiques_filiales,
    retourne la chaine de caractères output contenant les statistiques
    par filiale puis à la fin les statistiques de l'entreprise que
    l'on affichera par la suite.
    """
    effectif_total = 0
    somme_totale = 0
    salaires_min = []
    salaires_max = []
    output = ""

    for filiale in stats:
        effectif_total += filiale["effectif"]
        somme_totale += filiale["somme_salaire"]
        salaires_min.append(filiale["salaireMin"])
        salaires_max.append(filiale["salaireMax"])

        titre = f"Entreprise: {filiale['nom']}"
        output += titre + "\n\n"

        for employe in filiale["listeEmployes"]:
            ligne = f"{employe['name']:<10} | {employe['job']:<15} | {'Salaire mensuel :':<18} {employe['salary']:>10,.2f}"
            output += ligne + "\n"

        output += f"\nStatistiques des salaires pour la filiale : {filiale['nom']} :\n"
        output += f"Salaire moyen: {filiale['salaireMoyen'] :,.2f}€\n"
        output += f"Salaire le plus élevé : {filiale['salaireMax']:,.2f}€\n\n"
        output += f"Salaire le plus bas : {filiale['salaireMin']:,.2f}€\n"
        output += '=' * 60 + '\n'


    if effectif_total:
        salaire_moyen_global = somme_totale / effectif_total
    else:
        salaire_moyen_global = 0.0

    output += "\nStatistiques globales de l'entreprise :\n"
    output += f"Salaire moyen : {salaire_moyen_global:,.2f}€\n"
    output += f"Salaire le plus élevé : {max(salaires_max):,.2f}€\n"
    output += f"Salaire le plus bas : {min(salaires_min):,.2f}€\n"
    output += '=' * 60 + '\n'

    return output

def export_excel(stats: List[Dict[str, Any]], fichier: str = "statistiques_filiales.xlsx") -> None:
    """Exporte les statistiques au format Excel avec ajustement automatique de la largeur des colonnes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Statistiques"

    # En-têtes des colonnes
    entetes = ["Filiale", "Nom", "Poste", "Salaire", "Info"]
    ws.append(entetes)


    for filiale in stats:
        for employe in filiale["listeEmployes"]:
            ws.append([
                filiale["nom"],
                employe["name"],
                employe["job"],
                employe["salary"],
                employe["info"]
            ])

    # Pour chaque colonne
    for col in range(1, len(entetes) + 1):
        column = get_column_letter(col)
        max_length = 0
        # On cherche la longueur maximale pour chaque ligne
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        adjusted_width = (max_length + 3)
        # On ajuste la largeur avec +3 pour que le texte ne soit pas colé
        ws.column_dimensions[column].width = adjusted_width


    wb.save(fichier)
    print(f"Fichier Excel '{fichier}' créé avec succès.")


def export_csv(stats: List[Dict[str, Any]], fichier: str = "statistiques_filiales.csv") -> None:
    """Exporte les statistiques au format CSV."""
    with open(fichier, mode='w', newline='', encoding='utf-8') as f_csv:
        writer = csv.writer(f_csv, delimiter=';')
        entetes = ["Filiale", "Nom", "Poste", "Salaire", "Info"]
        writer.writerow(entetes)

        for filiale in stats:
            for employe in filiale["listeEmployes"]:
                writer.writerow([
                    filiale["nom"],
                    employe["name"],
                    employe["job"],
                    employe["salary"],
                    employe["info"]
                ])
    print(f"Fichier CSV '{fichier}' créé avec succès.")

def lancer_affichage(fichier):
    """
    charge le fichier json dans donnees
    charge statistiques_filiales(donnees) dans stats
    affiche la chaine ch_out = affichage(stats) dans le terminal, puis la retourne

    Args:
        fichier (str, optional): _description_. Defaults to './../data/employes-data.json'.
    """
    donnees = importation_des_donnees(fichier)
    stats = statistiques_filiales(donnees)
    ch_out = affichage(stats)
    print(ch_out)
    return ch_out

def export_salaries(lStats: List[Dict[str, Any]], filename: str, fmt: str = "csv", separator: str = ";") -> None:
    """
    Exporte les statistiques salariales sous format CSV ou Excel.

    Cette fonction génère un fichier contenant des informations sur les employés, ainsi que des statistiques globales et par filiale,
    incluant le salaire moyen, minimum, maximum, et la somme totale des salaires.

    Les données peuvent être exportées au format CSV ou Excel, selon la valeur du paramètre `fmt`. Les colonnes sont automatiquement ajustées
    dans le fichier Excel pour s'adapter aux données.

    Paramètres:
    -----------
    lStats : List[Dict[str, Any]]
        Liste des statistiques des filiales, où chaque filiale est représentée par un dictionnaire contenant des informations
        sur les employés (nom, poste, salaire, etc.) et des statistiques liées aux salaires.

    filename : str
        Le chemin et le nom du fichier dans lequel les données seront exportées.

    fmt : str, optionnel, par défaut "csv"
        Le format du fichier à générer, soit "csv" pour un fichier CSV, soit "xlsx" pour un fichier Excel.

    separator : str, optionnel, par défaut ";"
        Le séparateur à utiliser dans le fichier CSV (par exemple, ";" ou ",").

    Levée d'exceptions:
    -------------------
    ValueError :
        Si le format `fmt` n'est pas "csv" ou "xlsx".

    FileNotFoundError :
        Si le chemin du fichier spécifié est inaccessible en écriture.

    Exemple:
    --------
    export_salaries(lStats, "salaires.xlsx", fmt="xlsx")
    export_salaries(lStats, "salaires.csv", fmt="csv", separator=",")
    """
    if fmt not in ("csv", "xlsx"):
        raise ValueError("Format non supporté. Utilisez 'csv' ou 'xlsx'.")

    if not os.access(os.path.dirname(filename) or ".", os.W_OK):
        raise FileNotFoundError(f"Chemin inaccessible : {filename}")

    # --- Détail des employés ---
    employes = []
    total_employes = 0
    total_salaire = 0
    salaire_min = float('inf')
    salaire_max = float('-inf')

    for filiale in lStats:
        nb = filiale["effectif"]
        somme = filiale["somme_salaire"]
        total_employes += nb
        total_salaire += somme

        if filiale["salaireMin"] < salaire_min:
            salaire_min = filiale["salaireMin"]
        if filiale["salaireMax"] > salaire_max:
            salaire_max = filiale["salaireMax"]

        for emp in filiale["listeEmployes"]:
            employes.append({
                "Filiale": filiale["nom"],
                "Nom": emp["name"],
                "Poste": emp["job"],
                "Salaire mensuel": emp["salary"],
                "Heures sup": emp["info"]
            })

    salaire_moyen = total_salaire / total_employes if total_employes else 0

    if fmt == "csv":
        with open(filename, mode="w", newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=separator)

            # Données
            writer.writerow(["Filiale", "Nom", "Poste", "Salaire mensuel", "Heures sup"])
            for emp in employes:
                writer.writerow(emp.values())

            # Statistiques globales
            writer.writerow([])
            writer.writerow(["Statistiques globales"])
            writer.writerow(["Effectif total", total_employes])
            writer.writerow(["Salaire moyen", round(salaire_moyen, 2)])
            writer.writerow(["Salaire minimum", salaire_min])
            writer.writerow(["Salaire maximum", salaire_max])
            writer.writerow(["Somme totale des salaires", total_salaire])

            # Statistiques par filiale
            writer.writerow([])
            writer.writerow(["Statistiques par filiale"])
            writer.writerow(["Filiale", "Effectif", "Salaire moyen", "Salaire min", "Salaire max", "Somme salaires"])
            for f in lStats:
                writer.writerow([
                    f["nom"],
                    f["effectif"],
                    round(f["salaireMoyen"], 2),
                    f["salaireMin"],
                    f["salaireMax"],
                    f["somme_salaire"]
                ])

        print(f"Fichier CSV '{filename}' créé avec succès.")

    elif fmt == "xlsx":
        wb = Workbook()

        # Données
        ws1 = wb.active
        ws1.title = "Données"
        entetes = ["Filiale", "Nom", "Poste", "Salaire mensuel", "Heures sup"]
        ws1.append(entetes)
        for cell in ws1[1]:
            cell.font = Font(bold=True)
        for emp in employes:
            ws1.append(list(emp.values()))

        # Statistiques globales
        ws2 = wb.create_sheet("Stats Globales")
        ws2.append(["Statistiques globales"])
        ws2.append(["Effectif total", total_employes])
        ws2.append(["Salaire moyen", round(salaire_moyen, 2)])
        ws2.append(["Salaire minimum", salaire_min])
        ws2.append(["Salaire maximum", salaire_max])
        ws2.append(["Somme totale des salaires", total_salaire])
        for cell in ws2[1]:
            cell.font = Font(bold=True)

        # Statistiques par filiale
        ws3 = wb.create_sheet("Stats par Filiale")
        entetes_stats = ["Filiale", "Effectif", "Salaire moyen", "Salaire min", "Salaire max", "Somme salaires"]
        ws3.append(entetes_stats)
        for cell in ws3[1]:
            cell.font = Font(bold=True)

        for f in lStats:
            ws3.append([
                f["nom"],
                f["effectif"],
                round(f["salaireMoyen"], 2),
                f["salaireMin"],
                f["salaireMax"],
                f["somme_salaire"]
            ])

        # Largeur auto
        for ws in [ws1, ws2, ws3]:
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 2

        wb.save(filename)
        print(f"Fichier Excel '{filename}' créé avec succès.")



def lancer_tout(fichier):
    """
    charge le fichier json dans donnees
    charge statistiques_filiales(donnees) dans stats
    affiche la chaine ch_out = affichage(stats) dans le terminal
    exporte les donnees de stats dans les fichiers : "statistiques.xlsx" et "statistiques.csv"
    Args:
        fichier (str, optional): _description_. Defaults to './../data/employes-data.json'.
    """
    chemin0 = os.path.join(os.path.dirname(__file__), '..', 'data', 'employes-data.json')
    donnees = importation_des_donnees(chemin0)
    stats = statistiques_filiales(donnees)
    ch_out = affichage(stats)
#   print('\n' + '*' * 60 + '\n')
    print(ch_out)
    export_excel(stats, fichier="statistiques.xlsx")
    export_csv(stats, fichier="statistiques.csv")

    # Export CSV avec séparateur ;
    export_salaries(stats, "rapport_salaires.csv", fmt="csv", separator=';')

    # Export Excel avec mise en forme
    export_salaries(stats, "rapport_salaires.xlsx", fmt="xlsx")

def main():
    fichier = './../data/employes-data.json'  # Chemin du fichier JSON
    # Appeler la fonction lancer_tout pour afficher et exporter les données
    lancer_tout(fichier)

if __name__ == "__main__":
    main()
